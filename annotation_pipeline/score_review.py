"""Считает качество авторазметки по заполненному review_sample.xlsx"""
import sys, collections
from openpyxl import load_workbook

path = sys.argv[1] if len(sys.argv) > 1 else "review_sample.xlsx"
wb = load_workbook(path, data_only=True)
ws = wb["Разметка"]

rows = list(ws.iter_rows(min_row=2, values_only=True))
def g(r, i): 
    v = r[i]; 
    return ("" if v is None else str(v).strip())

SENTID, TEXT, TID, FORM, UP, LEM, MORPH, COR_UP, COR_LEM, COR_MORPH = range(10)
n = len(rows)
upos_err = 0
lem_err = 0
case_total = case_err = case_missed = 0
tense_total = tense_err = tense_missed = 0
err_by_pos = collections.Counter()
examples = collections.defaultdict(list)

for r in rows:
    if not g(r, FORM):
        continue
    auto_up = g(r, UP)
    auto_morph = g(r, MORPH)
    auto_lemm = g(r, LEM)
    # UPOS
    if g(r, COR_UP) and g(r, COR_UP) != auto_up:
        upos_err += 1; err_by_pos[auto_up] += 1
        if len(examples["UPOS"]) < 8:
            examples["UPOS"].append(f"{g(r,FORM)}: {auto_up} -> {g(r,COR_UP)}")
    # лемма
    if g(r, COR_LEM) and g(r, COR_LEM) != auto_lemm:
        lem_err += 1
        if len(examples["лемма"]) < 8:
            examples["лемма"].append(f"{g(r,FORM)}: {g(r,LEM)} -> {g(r,COR_LEM)}")
    # пока без проверки падежей и времени, разметка ненадежная
    # # падеж
    # if auto_morph.startswith("Case="):
    #     case_total += 1
    #     if g(r, COR_CASE):
    #         case_err += 1
    #         if len(examples["падеж"]) < 8:
    #             examples["падеж"].append(f"{g(r,FORM)}: {auto_morph} -> {g(r,COR_CASE)}")
    # elif g(r, COR_CASE):
    #     case_missed += 1
    # # время
    # if auto_morph.startswith("Tense="):
    #     tense_total += 1
    #     if g(r, COR_TENSE):
    #         tense_err += 1
    #         if len(examples["время"]) < 8:
    #             examples["время"].append(f"{g(r,FORM)}: {auto_morph} -> {g(r,COR_TENSE)}")
    # elif g(r, COR_TENSE):
    #     tense_missed += 1

def pct(ok, tot): return f"{100*ok/tot:.1f}%" if tot else "—"

print("="*52)
print(f"ПРОВЕРЕНО ТОКЕНОВ: {n}")
print("="*52)
print(f"UPOS (часть речи): точность {pct(n-upos_err, n)} ошибок {upos_err}/{n}")
print(f"Лемма: точность {pct(n-lem_err, n)} ошибок {lem_err}/{n}")
# print(f"Падеж (где размечен):точность {pct(case_total-case_err, case_total)} ошибок {case_err}/{case_total}; пропущено падежей: {case_missed}")
# print(f"Время (где размечено):точность {pct(tense_total-tense_err, tense_total)} ошибок {tense_err}/{tense_total}; пропущено времён: {tense_missed}")
print()
if err_by_pos:
    print("Ошибки UPOS по исходной метке машины:")
    for p, c in err_by_pos.most_common():
        print(f" {p}: {c}")
    print()
for k, exs in examples.items():
    if exs:
        print(f"Примеры ошибок ({k}):")
        for e in exs: print("   ", e)
        print()