"""Из авторазмеченного корпуса —> таблица для носителя"""
import os
import random
from conllu import parse_incr
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

random.seed(42)
_DIR = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(_DIR, "corpus_annotated.conllu")
N_SENT = 120

# падежи
CASES = ["Nom","Gen","Dat","All","Sup","Abl","Equ","Ine","Com","Dir","Voc","Ade"]
# времена глаголов
TENSES = ["Pres","Past","Fut"]
UPOS = ["NOUN","PROPN","VERB","ADJ","ADV","PRON","NUM","ADP","CCONJ","SCONJ","PART","INTJ","DET","AUX","X","PUNCT"]

def morph_of(t):
    fe = t["feats"] or {}
    if "Case" in fe: return "Case=" + fe["Case"]
    if "Tense" in fe: return "Tense=" + fe["Tense"]
    if "Mood" in fe: return "Mood=" + fe["Mood"]
    return ""

# отбор кандидатов
cands = []
with open(SRC, encoding="utf-8") as f:
    for s in parse_incr(f):
        toks = [t for t in s if isinstance(t["id"], int)]
        if not (5 <= len(toks) <= 25):
            continue
        text = s.metadata.get("text", "")
        if "ӕ" not in text and "Ӕ" not in text: # осетинский фильтр (в текстах встречаются и другие языки)
            continue
        x_share = sum(1 for t in toks if t["upos"] == "X") / len(toks)
        if x_share > 0.30:
            continue
        cands.append(s)

print("кандидатов:", len(cands))
sample = random.sample(cands, min(N_SENT, len(cands)))

# собираем таблицу
wb = Workbook()
ws = wb.active
ws.title = "Разметка"
header = ["sent_id","предложение","#","форма","авто_UPOS","авто_лемма","авто_морф", "верный UPOS","верная лемма","верный падеж","верное время","заметка"]
ws.append(header)

hfont = Font(bold=True, color="FFFFFF", name="Arial")
hfill = PatternFill("solid", start_color="2F5496")
for c in ws[1]:
    c.font = hfont; c.fill = hfill; c.alignment = Alignment(vertical="center", wrap_text=True)

sent_fill = PatternFill("solid", start_color="E8EEF7")
punct_fill = PatternFill("solid", start_color="F2F2F2")
thin = Side(style="thin", color="D9D9D9")
border = Border(bottom=thin)

row = 2
for s in sample:
    toks = [t for t in s if isinstance(t["id"], int)]
    text = s.metadata.get("text","")
    first = True
    for t in toks:
        # пунктуацию скипаем
        if t["upos"] == "PUNCT":
            continue 
        ws.cell(row, 1, s.metadata.get("sent_id",""))
        ws.cell(row, 2, text if first else "")
        ws.cell(row, 3, t["id"])
        ws.cell(row, 4, t["form"])
        ws.cell(row, 5, t["upos"])
        ws.cell(row, 6, t["lemma"])
        ws.cell(row, 7, morph_of(t))
        for col in range(1, 14):
            ws.cell(row, col).border = border
            ws.cell(row, col).font = Font(name="Arial", size=10)
        if first:
            for col in range(1, 14):
                ws.cell(row, col).fill = sent_fill
        row += 1
        first = False
    row += 0

# списки для фильтрации
dv_up = DataValidation(type="list", formula1='"%s"' % ",".join(UPOS), allow_blank=True)
dv_ca = DataValidation(type="list", formula1='"%s"' % ",".join(CASES), allow_blank=True)
dv_te = DataValidation(type="list", formula1='"%s"' % ",".join(TENSES), allow_blank=True)
for dv in (dv_up, dv_ca, dv_te):
    ws.add_data_validation(dv)
last = row - 1
dv_up.add(f"H2:H{last}")
dv_ca.add(f"I2:I{last}")
dv_te.add(f"K2:K{last}")

widths = [16, 40, 4, 16, 11, 16, 13, 9, 12, 13, 12, 12, 24]
for i,w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64+i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:M{last}"

out = os.path.join(_DIR, "review_sample.xlsx")
wb.save(out)
print("предложений в выборке:", len(sample), "строк:", last-1)
